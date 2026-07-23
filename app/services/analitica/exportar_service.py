from datetime import datetime
from io import BytesIO
from typing import Optional
from xml.sax.saxutils import escape

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image


def _label(key: str) -> str:
    return key.replace("_", " ").title()


def _valor_str(value) -> str:
    """Convierte cualquier valor a string legible para una celda."""
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, dict):
        return "\n".join(f"{_label(k)}: {v}" for k, v in value.items())
    if isinstance(value, list):
        if not value:
            return "—"
        if isinstance(value[0], dict):
            partes = []
            for item in value:
                partes.append("  •  " + "  |  ".join(f"{_label(k)}: {v}" for k, v in item.items()))
            return "\n".join(partes)
        return "\n".join(f"• {v}" for v in value)
    return str(value)


def _rango_fechas_str(fecha_inicio: Optional[str], fecha_fin: Optional[str]) -> Optional[str]:
    if fecha_inicio and fecha_fin:
        return f"Período: {fecha_inicio}  →  {fecha_fin}"
    return None


def generar_pdf(
    titulo: str,
    transcripcion: str,
    datos: dict,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("AutoSOS — Reporte Media Voz", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(titulo, styles["Heading2"]))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))

    rango = _rango_fechas_str(fecha_inicio, fecha_fin)
    if rango:
        elements.append(Paragraph(rango, styles["Normal"]))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>Consulta:</b> {transcripcion}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    table_data = [["Métrica", "Valor"]]
    for key, value in datos.items():
        table_data.append([_label(key), _valor_str(value)])

    col_widths = [7 * cm, 11 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#e8eaf6")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9e9e9e")),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_excel(
    titulo: str,
    transcripcion: str,
    datos: dict,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    azul = "1a237e"
    azul_claro = "e8eaf6"
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Fila 1 — título
    ws.merge_cells("A1:B1")
    ws["A1"] = "AutoSOS — Reporte Media Voz"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2 — subtítulo
    ws.merge_cells("A2:B2")
    ws["A2"] = titulo
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Fila 3 — fecha generación
    ws.merge_cells("A3:B3")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(italic=True, size=10, color="555555")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Fila 4 — rango de fechas (si existe) o consulta
    fila_consulta = 4
    rango = _rango_fechas_str(fecha_inicio, fecha_fin)
    if rango:
        ws.merge_cells("A4:B4")
        ws["A4"] = rango
        ws["A4"].font = Font(italic=True, size=10, color="1a237e")
        ws["A4"].alignment = Alignment(horizontal="center")
        fila_consulta = 5

    ws.merge_cells(f"A{fila_consulta}:B{fila_consulta}")
    ws[f"A{fila_consulta}"] = f"Consulta: {transcripcion}"
    ws[f"A{fila_consulta}"].font = Font(size=10)
    ws[f"A{fila_consulta}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[fila_consulta].height = 36

    # Encabezados de tabla
    fila_header = fila_consulta + 1
    ws.cell(row=fila_header, column=1, value="Métrica").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=fila_header, column=2, value="Valor").font = Font(bold=True, color="FFFFFF")
    for col in [1, 2]:
        c = ws.cell(row=fila_header, column=col)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center")
        c.border = borde

    # Datos
    for i, (key, value) in enumerate(datos.items(), start=fila_header + 1):
        celda_label = ws.cell(row=i, column=1, value=_label(key))
        celda_valor = ws.cell(row=i, column=2, value=_valor_str(value))
        celda_label.border = borde
        celda_valor.border = borde
        celda_valor.alignment = Alignment(wrap_text=True, vertical="top")
        celda_label.alignment = Alignment(vertical="top")
        if i % 2 == 0:
            celda_label.fill = PatternFill("solid", fgColor=azul_claro)
            celda_valor.fill = PatternFill("solid", fgColor=azul_claro)
        # Altura dinámica para listas largas (técnicos)
        if isinstance(value, list) and len(value) > 1:
            ws.row_dimensions[i].height = max(15 * len(value), 30)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Reporte de siniestro (para aseguradoras)
# ---------------------------------------------------------------------------

def _fecha_str(value) -> str:
    if value is None:
        return "—"
    try:
        return value.strftime("%d/%m/%Y %H:%M")
    except AttributeError:
        return str(value)


_ESTILO_CELDA_KV = ParagraphStyle(
    "CeldaKV", fontName="Helvetica", fontSize=9.5, leading=12.5, wordWrap="CJK"
)
_ESTILO_CELDA_KV_LABEL = ParagraphStyle(
    "CeldaKVLabel", parent=_ESTILO_CELDA_KV, fontName="Helvetica-Bold"
)


def _celda(texto, negrita: bool = False) -> Paragraph:
    texto = "—" if texto is None or texto == "" else str(texto)
    estilo = _ESTILO_CELDA_KV_LABEL if negrita else _ESTILO_CELDA_KV
    return Paragraph(escape(texto).replace("\n", "<br/>"), estilo)


def _tabla_kv(pares: list[tuple[str, str]]) -> Table:
    tabla = Table(
        [[_celda(k, negrita=True), _celda(v)] for k, v in pares],
        colWidths=[5 * cm, 11 * cm],
    )
    tabla.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c5cae9")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8eaf6")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return tabla


def _descargar_imagen_flowable(url: str, ancho_max: float = 9 * cm, alto_max: float = 10 * cm) -> Optional[Image]:
    try:
        respuesta = requests.get(url, timeout=10)
        respuesta.raise_for_status()
        contenido = respuesta.content
        pil_img = PILImage.open(BytesIO(contenido))
        ancho_px, alto_px = pil_img.size
        ratio = ancho_px / alto_px if alto_px else 1
        ancho, alto = ancho_max, ancho_max / ratio
        if alto > alto_max:
            ancho, alto = alto_max * ratio, alto_max
        return Image(BytesIO(contenido), width=ancho, height=alto)
    except Exception:
        return None


def generar_pdf_siniestro(datos: dict) -> BytesIO:
    """Genera el PDF de 'Reporte de Siniestro' con todo lo recopilado de un incidente,
    listo para adjuntarse a un trámite ante una aseguradora."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    incidente = datos.get("incidente") or {}
    cliente = datos.get("cliente")
    vehiculo = datos.get("vehiculo")
    analisis = datos.get("analisis")
    taller = datos.get("taller_atendio")
    orden = datos.get("orden_servicio") or {}
    transaccion = datos.get("transaccion")
    evidencias = datos.get("evidencias") or []

    elements.append(Paragraph("AutoSOS — Reporte de Siniestro", styles["Title"]))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Código de siniestro: <b>{datos.get('codigo_siniestro', '—')}</b>", styles["Normal"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("1. Datos del Incidente", styles["Heading2"]))
    elements.append(
        _tabla_kv(
            [
                ("N° de incidente", str(incidente.get("id", "—"))),
                ("Fecha y hora", _fecha_str(incidente.get("fecha_hora"))),
                ("Estado", _valor_str(incidente.get("estado"))),
                ("Prioridad", _valor_str(incidente.get("prioridad"))),
                ("Ubicación (lat, lon)", f"{incidente.get('latitud', '—')}, {incidente.get('longitud', '—')}"),
            ]
        )
    )
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("2. Datos del Asegurado", styles["Heading2"]))
    if cliente:
        elements.append(
            _tabla_kv(
                [
                    ("Nombre", cliente.get("nombre", "—")),
                    ("Teléfono", cliente.get("telefono", "—")),
                    ("Email", cliente.get("email") or "—"),
                ]
            )
        )
    else:
        elements.append(Paragraph("No se encontró información del cliente.", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("3. Datos del Vehículo", styles["Heading2"]))
    if vehiculo:
        elements.append(
            _tabla_kv(
                [
                    ("Placa", vehiculo.get("placa", "—")),
                    ("Modelo", vehiculo.get("modelo", "—")),
                    ("Color", vehiculo.get("color", "—")),
                ]
            )
        )
    else:
        elements.append(Paragraph("No se encontró vehículo registrado para este cliente.", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("4. Diagnóstico (Análisis por IA)", styles["Heading2"]))
    if analisis:
        elements.append(
            _tabla_kv(
                [
                    ("Categoría del problema", analisis.get("categoria_problema", "—")),
                    ("Daños identificados", analisis.get("danios_identificados", "—")),
                    ("Resumen", analisis.get("resumen_estructurado", "—")),
                    ("Transcripción del reporte", analisis.get("transcripcion_audio", "—")),
                ]
            )
        )
    else:
        elements.append(Paragraph("Este incidente no cuenta con análisis de IA registrado.", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("5. Taller que Atendió", styles["Heading2"]))
    if taller:
        elements.append(
            _tabla_kv(
                [
                    ("Nombre", taller.get("nombre", "—")),
                    ("Teléfono", taller.get("telefono", "—")),
                    ("Dirección", taller.get("direccion", "—")),
                ]
            )
        )
    else:
        elements.append(Paragraph("El incidente no fue atendido por ningún taller (sin asignación aceptada).", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("6. Servicios Realizados", styles["Heading2"]))
    detalles = orden.get("detalles") or []
    if detalles:
        tabla_data = [["Servicio", "Cant.", "P. Unitario", "Subtotal"]]
        total = 0.0
        for d in detalles:
            tabla_data.append(
                [
                    _celda(d.get("servicio_nombre")),
                    str(d.get("cantidad", "—")),
                    f"{d.get('precio_unitario', 0):,.2f}",
                    f"{d.get('subtotal', 0):,.2f}",
                ]
            )
            total += d.get("subtotal") or 0
        tabla_data.append(["", "", "Total", f"{total:,.2f}"])
        tabla = Table(tabla_data, colWidths=[7 * cm, 2 * cm, 3.5 * cm, 3.5 * cm])
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9e9e9e")),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(tabla)
    else:
        elements.append(Paragraph("No hay servicios/costos registrados para este incidente.", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("7. Estado de Pago", styles["Heading2"]))
    if transaccion:
        elements.append(
            _tabla_kv(
                [
                    ("Monto cobrado", f"{transaccion.get('monto_cobrado', 0):,.2f}"),
                    ("Método de pago", _valor_str(transaccion.get("metodo_pago"))),
                    ("Estado", _valor_str(transaccion.get("estado"))),
                    ("Fecha", _fecha_str(transaccion.get("fecha_hora"))),
                ]
            )
        )
    else:
        elements.append(Paragraph("No se registró transacción de pago.", styles["Normal"]))
    elements.append(Spacer(1, 12))

    fotos = [e for e in evidencias if str(e.get("tipo")) == "Foto" and e.get("url")]
    otras = [e for e in evidencias if str(e.get("tipo")) != "Foto"]

    if fotos:
        elements.append(Paragraph("8. Evidencia Fotográfica", styles["Heading2"]))
        for e in fotos:
            elements.append(Paragraph(f"Evidencia #{e.get('id')} — {_fecha_str(e.get('fecha_subida'))}", styles["Normal"]))
            elements.append(Spacer(1, 4))
            imagen = _descargar_imagen_flowable(e["url"])
            elements.append(imagen if imagen else Paragraph("(No se pudo descargar la imagen)", styles["Italic"]))
            elements.append(Spacer(1, 10))

    if otras:
        elements.append(Paragraph("Otras evidencias", styles["Heading3"]))
        for e in otras:
            elements.append(Paragraph(f"• {_valor_str(e.get('tipo'))}: {e.get('url', '—')}", styles["Normal"]))
        elements.append(Spacer(1, 10))

    elements.append(Spacer(1, 16))
    elements.append(
        Paragraph(
            "Este documento fue generado automáticamente por AutoSOS a partir de los registros del incidente "
            "y sirve como respaldo informativo para el trámite del siniestro ante la aseguradora. "
            "No constituye una pericia técnica oficial.",
            styles["Italic"],
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer